import os
from openpyxl import load_workbook, Workbook
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    abort
)
from utils.auth import login_required
from flask import session
from sqlalchemy import or_
from flask import request, jsonify
from models.field_correction import FieldCorrection
from models.dataset_entry import DatasetEntry
from models import db
from models.project import Project
from datetime import datetime
project = Blueprint("project", __name__)

@project.route("/project/<int:project_id>/dashboard")
@login_required


def project_dashboard(project_id):

    project = Project.query.get_or_404(project_id)

    total_records = DatasetEntry.query.filter_by(
        project_id=project.id
    ).count()

    total_topics = db.session.query(
        DatasetEntry.topic_id
    ).filter_by(
        project_id=project.id
    ).distinct().count()
    
    not_started = DatasetEntry.query.filter_by(
    project_id=project.id,
    attempted=False
    ).count()

    english_only = DatasetEntry.query.filter(
    DatasetEntry.project_id == project.id,
    DatasetEntry.attempted == True,
    or_(
        DatasetEntry.question_ml.is_(None),
        DatasetEntry.question_ml == ""
    )
    ).count()

    completed = DatasetEntry.query.filter(
        DatasetEntry.project_id == project.id,
        DatasetEntry.question_ml.isnot(None),
        DatasetEntry.question_ml != ""
    ).count()

    corrections = DatasetEntry.query.filter(
        DatasetEntry.project_id == project.id,
        DatasetEntry.correction_note.isnot(None),
        DatasetEntry.correction_note != ""
    ).count()
    progress = 0
    draft_count = DatasetEntry.query.filter_by(
    project_id=project.id,
    status=DatasetEntry.STATUS_DRAFT
    ).count()

    submitted_count = DatasetEntry.query.filter_by(
        project_id=project.id,
        status=DatasetEntry.STATUS_SUBMITTED
    ).count()

    corrections_requested_count = DatasetEntry.query.filter_by(
        project_id=project.id,
        status=DatasetEntry.STATUS_CORRECTIONS
    ).count()

    approved_count = DatasetEntry.query.filter_by(
        project_id=project.id,
        status=DatasetEntry.STATUS_APPROVED
    ).count()
    if total_records > 0:
        progress = round((completed / total_records) * 100)
        
    role = session.get("role")

    template = (
        "reviewer_dashboard.html"
        if role == "REVIEWER"
        else "project_dashboard.html"
    )

    return render_template(
        template,
        project=project,

        total_records=total_records,
        total_topics=total_topics,

        not_started=not_started,
        english_only=english_only,
        completed=completed,
        corrections=corrections,

        progress=progress,

        draft_count=draft_count,
        submitted_count=submitted_count,
        corrections_requested_count=corrections_requested_count,
        approved_count=approved_count,

        role=role
    )
    
@project.route("/project/<int:project_id>/dataset")
def view_dataset(project_id):

    project = Project.query.get_or_404(project_id)

    search = request.args.get("search", "").strip()
    
    topic = request.args.get("topic", "")

    query = DatasetEntry.query.filter_by(project_id=project.id)
    
    query = query.order_by(
        DatasetEntry.topic_id.asc(),
        DatasetEntry.practice_question_id.asc()
    )
    
    if topic:
        query = query.filter(
            DatasetEntry.topic_id == topic
        )

    if search:

        query = query.filter(

            or_(

                DatasetEntry.topic_en.ilike(f"%{search}%"),
                DatasetEntry.topic_ml.ilike(f"%{search}%"),

                DatasetEntry.sub_topic_en.ilike(f"%{search}%"),
                DatasetEntry.sub_topic_ml.ilike(f"%{search}%"),

                DatasetEntry.practice_question_id.ilike(f"%{search}%"),

                DatasetEntry.question_en.ilike(f"%{search}%")

            )

        )
        
    topics = (
        db.session.query(DatasetEntry.topic_id)
        .filter_by(project_id=project.id)
        .distinct()
        .order_by(DatasetEntry.topic_id)
        .all()
    )
        

    page = request.args.get("page", 1, type=int)

    dataset = query.paginate(
        page=page,
        per_page=20,
        error_out=False
    )
    
    total_records = DatasetEntry.query.filter_by(
        project_id=project.id
    ).count()
    
    role = session.get("role")
    print("CURRENT ROLE =", repr(role))

    return render_template(
        "view_dataset.html",
        project=project,
        dataset=dataset,
        search=search,
        topic=topic,
        topics=topics,
        total_records=total_records,
        role=role,
    )
    
    
@project.route("/question/<int:entry_id>")
def view_question(entry_id):

    entry = DatasetEntry.query.get_or_404(entry_id)

    return render_template(
        "view_question.html",
        entry=entry
    )


@project.route("/question/<int:entry_id>/edit", methods=["GET", "POST"])
def edit_question(entry_id):

    entry = DatasetEntry.query.get_or_404(entry_id)
    
    print("\n========== EDIT PAGE DEBUG ==========")
    print("DB ID:", entry.id)
    print("Practice ID:", entry.practice_question_id)
    print("Question:", repr(entry.question_en))
    print("Option A:", repr(entry.option_a_en))
    print("Option B:", repr(entry.option_b_en))
    print("Option C:", repr(entry.option_c_en))
    print("Correct Answer:", repr(entry.correct_answer_en))
    print("=====================================\n")

    shared_fields_locked = (
        entry.practice_question_id
        and not entry.practice_question_id.endswith("_Q1")
    )

    if request.method == "POST":
        action = request.form.get("action", "save")
        
        if entry.status in [
            DatasetEntry.STATUS_SUBMITTED,
            DatasetEntry.STATUS_APPROVED
        ]:
            flash("This question is currently locked for editing.", "warning")

            return redirect(
                url_for(
                    "project.edit_question",
                    entry_id=entry.id
                )
            )
                
        
    
        
        

    # -------------------------------
    # Topic
    # -------------------------------
        # -------------------------------
        # Shared Fields
        # Only Q1 can update these
        # -------------------------------

        if not shared_fields_locked:

            # Topic
            entry.topic_en = request.form.get("topic_en")
            entry.topic_ml = request.form.get("topic_ml")

            # Sub Topic
            selected_subtopic = request.form.get("sub_topic_en")

            if selected_subtopic == "__NEW__":
                entry.sub_topic_en = request.form.get("new_sub_topic")
            else:
                entry.sub_topic_en = selected_subtopic

            entry.sub_topic_ml = request.form.get("sub_topic_ml")

            # Scenario
            entry.scenario_en = request.form.get("scenario_en")
            entry.scenario_ml = request.form.get("scenario_ml")

            entry.scenario_explanation_en = request.form.get(
                "scenario_explanation_en"
            )
            entry.scenario_explanation_ml = request.form.get(
                "scenario_explanation_ml"
            )

            entry.memory_shortcut_en = request.form.get(
                "memory_shortcut_en"
            )
            entry.memory_shortcut_ml = request.form.get(
                "memory_shortcut_ml"
            )

            entry.applies_when_en = request.form.get(
                "applies_when_en"
            )
            entry.applies_when_ml = request.form.get(
                "applies_when_ml"
            )

            entry.not_applies_when_en = request.form.get(
                "not_applies_when_en"
            )
            entry.not_applies_when_ml = request.form.get(
                "not_applies_when_ml"
            )

        # -------------------------------
        # Question
        # -------------------------------
        entry.question_en = request.form.get("question_en")
        entry.question_ml = request.form.get("question_ml")
        
        entry.option_a_en = request.form.get("option_a_en")
        entry.option_a_ml = request.form.get("option_a_ml")

        entry.option_b_en = request.form.get("option_b_en")
        entry.option_b_ml = request.form.get("option_b_ml")

        entry.option_c_en = request.form.get("option_c_en")
        entry.option_c_ml = request.form.get("option_c_ml")

        entry.correct_answer_letter = request.form.get("correct_answer_letter")

        entry.correct_answer_en = request.form.get("correct_answer_en")
        entry.correct_answer_ml = request.form.get("correct_answer_ml")

        entry.explanation_en = request.form.get("explanation_en")
        entry.explanation_ml = request.form.get("explanation_ml")

        entry.wrong_answer_tip_en = request.form.get("wrong_answer_tip_en")
        entry.wrong_answer_tip_ml = request.form.get("wrong_answer_tip_ml")

    # -------------------------------
    # Other Fields
    # -------------------------------
        entry.question_image_ref = request.form.get("question_image_ref")

        entry.correction_note = request.form.get("correction_note")

        entry.english_completed = "english_completed" in request.form
        entry.malayalam_completed = "malayalam_completed" in request.form

        entry.attempted = True
        
        if action == "submit_review":
            entry.status = DatasetEntry.STATUS_SUBMITTED
            entry.submitted_at = datetime.utcnow()
            
        elif action == "approve":
            entry.status = DatasetEntry.STATUS_APPROVED
            entry.approved_at = datetime.utcnow()

        elif action == "corrections":
            entry.status = DatasetEntry.STATUS_CORRECTIONS

        else:
            # Save Changes
            if entry.status is None:
                entry.status = DatasetEntry.STATUS_DRAFT
                
                
        rich_content = dict(entry.rich_content or {})


        shared_field_names = {
            "topic_en",
            "topic_ml",
            "sub_topic_en",
            "sub_topic_ml",
            "scenario_en",
            "scenario_ml",
            "scenario_explanation_en",
            "scenario_explanation_ml",
            "memory_shortcut_en",
            "memory_shortcut_ml",
            "applies_when_en",
            "applies_when_ml",
            "not_applies_when_en",
            "not_applies_when_ml",
        }

        for key, value in request.form.items():

            if not key.startswith("rich_"):
                continue

            field_name = key[5:]

            # Q2+ cannot modify shared-field formatting
            if shared_fields_locked and field_name in shared_field_names:
                continue

            if value:
                rich_content[field_name] = value
            else:
                rich_content.pop(field_name, None)


        entry.rich_content = rich_content
        
        try:
            db.session.commit()

        except Exception as e:

            # Reset the failed SQLAlchemy transaction
            db.session.rollback()

            # Keep the actual error in the server logs
            print(
                f"QUESTION SAVE FAILED | "
                f"Entry ID: {entry.id} | "
                f"Error: {repr(e)}"
            )

            flash(
                "The question could not be saved due to a temporary "
                "server or database issue. Please try again.",
                "danger"
            )

            return redirect(
                url_for(
                    "project.edit_question",
                    entry_id=entry.id
                )
            )


        flash(f"Current Status: {entry.status}", "info")
        flash("Question updated successfully!", "success")

        return redirect(
            url_for(
                "project.edit_question",
                entry_id=entry.id,
                saved=1
            )
        )
    project = Project.query.get(entry.project_id)

    # -------------------------------------------------
    # Previous / Next navigation
    # Keep navigation inside the workflow queue
    # the user came from.
    # -------------------------------------------------

    from_status = request.args.get("from_status")

    status_map = {
        "Draft": DatasetEntry.STATUS_DRAFT,
        "Pending Review": DatasetEntry.STATUS_SUBMITTED,
        "Corrections Requested": DatasetEntry.STATUS_CORRECTIONS,
        "Approved": DatasetEntry.STATUS_APPROVED,
    }

    queue_status = status_map.get(from_status)


    if queue_status:

        # User entered this page from one of the
        # Dashboard workflow queues.
        previous_entry = DatasetEntry.query.filter(
            DatasetEntry.project_id == entry.project_id,
            DatasetEntry.status == queue_status,
            DatasetEntry.id < entry.id
        ).order_by(
            DatasetEntry.id.desc()
        ).first()

        next_entry = DatasetEntry.query.filter(
            DatasetEntry.project_id == entry.project_id,
            DatasetEntry.status == queue_status,
            DatasetEntry.id > entry.id
        ).order_by(
            DatasetEntry.id.asc()
        ).first()

    else:

        # Normal editing behaviour when the page
        # was NOT opened from a workflow queue.
        previous_entry = DatasetEntry.query.filter(
            DatasetEntry.project_id == entry.project_id,
            DatasetEntry.id < entry.id
        ).order_by(
            DatasetEntry.id.desc()
        ).first()

        next_entry = DatasetEntry.query.filter(
            DatasetEntry.project_id == entry.project_id,
            DatasetEntry.attempted == False,
            DatasetEntry.id > entry.id
        ).order_by(
            DatasetEntry.id.asc()
        ).first()

    topics = (
    db.session.query(DatasetEntry.topic_en)
    .distinct()
    .order_by(DatasetEntry.topic_en)
    .all()
    )

    topics = [t[0] for t in topics if t[0]]
    

    subtopics = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .distinct()
        .order_by(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .all()
    )
    
    
    topic_translations = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.topic_ml
        )
        .distinct()
        .all()
    )

    subtopic_translations = (
        db.session.query(
            DatasetEntry.sub_topic_en,
            DatasetEntry.sub_topic_ml
        )
        .distinct()
        .all()
    )
    
    role = session.get("role")
    team_editing_locked = (
        entry.status in [
            DatasetEntry.STATUS_SUBMITTED,
            DatasetEntry.STATUS_APPROVED
        ]
    )
    
    corrections = FieldCorrection.query.filter_by(
        entry_id=entry.id,
        is_active=True
    ).all()
    correction_count = len(corrections)

    correction_map = {}

    for correction in corrections:
        correction_map[correction.field_name] = correction.comment
        
    
    shared_entry = None

    if shared_fields_locked:

        q1_id = entry.practice_question_id.rsplit("_Q", 1)[0] + "_Q1"

        shared_entry = DatasetEntry.query.filter_by(
            project_id=entry.project_id,
            practice_question_id=q1_id
        ).first()
        
        
    return render_template(
        "edit_question.html",
        project=project,
        entry=entry,
        previous_entry=previous_entry,
        next_entry=next_entry,
        topics=topics,
        subtopics=subtopics,
        topic_translations=topic_translations,
        subtopic_translations=subtopic_translations,
        role=role,
        correction_map=correction_map,
        correction_count=correction_count,
        subtopic_details={},
        shared_fields_locked=shared_fields_locked,
        shared_entry=shared_entry,
        team_editing_locked=team_editing_locked,
    )
        
    

@project.route("/question/<int:entry_id>/delete", methods=["POST"])
def delete_question(entry_id):

    entry = DatasetEntry.query.get_or_404(entry_id)
    
    

    project_id = entry.project_id

    db.session.delete(entry)
    db.session.commit()

    flash("Question deleted successfully!", "success")

    return redirect(
        url_for(
            "project.view_dataset",
            project_id=project_id
        )
    )
    



@project.route("/create-project", methods=["GET", "POST"])
def create_project():

    if request.method == "POST":

        project_name = request.form.get("project_name")

        description = request.form.get("description")

        new_project = Project(
            project_name=project_name,
            description=description
        )

        db.session.add(new_project)

        db.session.commit()

        return redirect(url_for("workspace.workspace_page"))

    return render_template("create_project.html")

@project.route(
    "/project/<int:project_id>/import",
    methods=["GET","POST"]
)
def import_dataset(project_id):

    project = Project.query.get_or_404(project_id)

    if request.method == "POST":

        # ==========================================
        # IMPORT SETTINGS
        # ==========================================

        import_mode = request.form.get(
            "import_mode",
            "english_only"
        )

        existing_action = request.form.get(
            "existing_action",
            "new"
        )

        english_file = request.files.get(
            "english_dataset"
        )
        

        malayalam_file = request.files.get(
            "malayalam_dataset"
        )



        # ==========================================
        # VALIDATE SETTINGS
        # ==========================================

        if import_mode not in {
            "english_only",
            "english_malayalam",
            "malayalam_only"
        }:

            flash("Invalid import type.", "danger")
            return redirect(request.url)


        if existing_action not in {
            "new",
            "merge",
            "replace"
        }:

            flash("Invalid dataset action.", "danger")
            return redirect(request.url)


        # English is always required
        # ==========================================
        # VALIDATE FILES FOR IMPORT MODE
        # ==========================================

        if import_mode == "english_only":

            if not english_file or not english_file.filename:
                flash(
                    "Please select the English dataset.",
                    "danger"
                )
                return redirect(request.url)


        elif import_mode == "english_malayalam":

            if not english_file or not english_file.filename:
                flash(
                    "Please select the English dataset.",
                    "danger"
                )
                return redirect(request.url)

            if not malayalam_file or not malayalam_file.filename:
                flash(
                    "Please select the Malayalam dataset.",
                    "danger"
                )
                return redirect(request.url)


        elif import_mode == "malayalam_only":

            if not malayalam_file or not malayalam_file.filename:
                flash(
                    "Please select the Malayalam dataset.",
                    "danger"
                )
                return redirect(request.url)

            existing_count = DatasetEntry.query.filter_by(
                project_id=project.id
            ).count()

            if existing_count == 0:
                flash(
                    "Import the English dataset before importing Malayalam translations.",
                    "danger"
                )
                return redirect(request.url)


        # Malayalam required for dual import
        if import_mode == "english_malayalam":

            if (
                not malayalam_file
                or not malayalam_file.filename
            ):

                flash(
                    "Please select the Malayalam dataset.",
                    "danger"
                )

                return redirect(request.url)


        # ==========================================
        # PROJECT UPLOAD FOLDER
        # ==========================================

        upload_folder = os.path.join(
            "uploads",
            f"project_{project.id}"
        )

        os.makedirs(
            upload_folder,
            exist_ok=True
        )


        # ==========================================
        # SAVE ENGLISH FILE
        # ==========================================

        english_path = None

        if import_mode in {
            "english_only",
            "english_malayalam"
        }:

            english_path = os.path.join(
                upload_folder,
                "backbone_english.xlsx"
            )

            english_file.save(english_path)

        # ==========================================
        # SAVE MALAYALAM FILE IF PROVIDED
        # ==========================================

        malayalam_path = None

        if import_mode in {
            "english_malayalam",
            "malayalam_only"
        }:

            malayalam_path = os.path.join(
                upload_folder,
                "backbone_malayalam.xlsx"
            )

            malayalam_file.save(
                malayalam_path
            )
            
            
        # ==========================================
        # MALAYALAM-ONLY IMPORT
        # ==========================================

        if import_mode == "malayalam_only":

            try:
                ml_workbook = load_workbook(
                    malayalam_path,
                    data_only=True
                )

                ml_sheet = ml_workbook.active

                ml_headers = next(
                    ml_sheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        values_only=True
                    )
                )

            except Exception as e:

                print("Failed to open Malayalam workbook:", e)

                flash(
                    "Could not read the Malayalam Excel file.",
                    "danger"
                )

                return redirect(request.url)


            # Carry forward shared fields
            ml_current_topic = None
            ml_current_sub_topic = None
            ml_current_scenario = None
            ml_current_scenario_explanation = None
            ml_current_memory_shortcut = None
            ml_current_applies_when = None
            ml_current_not_applies_when = None

            matched_count = 0
            unmatched_count = 0

            # Load all existing project entries once
            project_entries = DatasetEntry.query.filter_by(
                project_id=project.id
            ).all()

            ml_existing_entries = {
                str(item.practice_question_id).strip(): item
                for item in project_entries
                if item.practice_question_id
            }
            

            for ml_row in ml_sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                ml_data = {
                    header: value
                    for header, value in zip(
                        ml_headers,
                        ml_row
                    )
                    if header is not None
                }

                practice_id = ml_data.get(
                    "Practice_question_id"
                )

                if not practice_id:
                    continue

                practice_id = str(practice_id).strip()


                # Remember shared Malayalam values
                if ml_data.get("Topic"):
                    ml_current_topic = ml_data["Topic"]

                if ml_data.get("Sub_topic"):
                    ml_current_sub_topic = ml_data["Sub_topic"]

                if ml_data.get("Scenario"):
                    ml_current_scenario = ml_data["Scenario"]

                if ml_data.get("Scenario_explanation"):
                    ml_current_scenario_explanation = (
                        ml_data["Scenario_explanation"]
                    )

                if ml_data.get("Memory_shortcut"):
                    ml_current_memory_shortcut = (
                        ml_data["Memory_shortcut"]
                    )

                if ml_data.get("Applies_when"):
                    ml_current_applies_when = (
                        ml_data["Applies_when"]
                    )

                if ml_data.get("Not_applies_when"):
                    ml_current_not_applies_when = (
                        ml_data["Not_applies_when"]
                    )


                # Find corresponding existing English entry
                existing_entry = ml_existing_entries.get(practice_id)


                if not existing_entry:

                    unmatched_count += 1

                    print(
                        "Malayalam ID not found:",
                        practice_id
                    )

                    continue


                # Fill Malayalam fields
                existing_entry.topic_ml = ml_current_topic
                existing_entry.sub_topic_ml = ml_current_sub_topic
                existing_entry.scenario_ml = ml_current_scenario

                existing_entry.scenario_explanation_ml = (
                    ml_current_scenario_explanation
                )

                existing_entry.memory_shortcut_ml = (
                    ml_current_memory_shortcut
                )

                existing_entry.applies_when_ml = (
                    ml_current_applies_when
                )

                existing_entry.not_applies_when_ml = (
                    ml_current_not_applies_when
                )

                existing_entry.question_ml = (
                    ml_data.get("Questions")
                )

                existing_entry.option_a_ml = (
                    ml_data.get("Option_A")
                )

                existing_entry.option_b_ml = (
                    ml_data.get("Option_B")
                )

                existing_entry.option_c_ml = (
                    ml_data.get("Option_C")
                )

                existing_entry.correct_answer_ml = (
                    ml_data.get("Correct_answer")
                )

                existing_entry.explanation_ml = (
                    ml_data.get("Explanation")
                )

                existing_entry.wrong_answer_tip_ml = (
                    ml_data.get("Wrong_Answer_Tip")
                )

                matched_count += 1



            try:

                db.session.commit()

            except Exception as e:

                db.session.rollback()

                print(
                    "MALAYALAM IMPORT FAILED:",
                    e
                )

                flash(
                    "Malayalam import failed. No changes were saved.",
                    "danger"
                )

                return redirect(request.url)


            print(
                "Malayalam import complete.",
                "Matched:",
                matched_count,
                "Unmatched:",
                unmatched_count
            )

            flash(
                f"Malayalam imported successfully. "
                f"{matched_count} entries matched, "
                f"{unmatched_count} unmatched.",
                "success"
            )

            return redirect(
                url_for(
                    "project.project_dashboard",
                    project_id=project.id
                )
            )


        # ==========================================
        # OPEN ENGLISH WORKBOOK
        # ==========================================

        try:

            workbook = load_workbook(
                english_path,
                data_only=True
            )

            sheet = workbook.active

        except Exception as e:

            print("Failed to open English workbook:", e)

            flash(
                "Could not read the English Excel file.",
                "danger"
            )

            return redirect(request.url)
        
        
        # ==========================================
        # READ MALAYALAM DATA
        # ==========================================

        malayalam_lookup = {}

        if import_mode == "english_malayalam":

            try:

                ml_workbook = load_workbook(
                    malayalam_path,
                    data_only=True
                )

                ml_sheet = ml_workbook.active

                ml_headers = next(
                    ml_sheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        values_only=True
                    )
                )

                # Carry forward shared Malayalam fields
                ml_current_topic = None
                ml_current_sub_topic = None
                ml_current_scenario = None
                ml_current_scenario_explanation = None
                ml_current_memory_shortcut = None
                ml_current_applies_when = None
                ml_current_not_applies_when = None

                for ml_row in ml_sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ):

                    ml_data = {
                        header: value
                        for header, value in zip(
                            ml_headers,
                            ml_row
                        )
                        if header is not None
                    }

                    practice_id = ml_data.get(
                        "Practice_question_id"
                    )

                    if not practice_id:
                        continue

                    # Shared fields may only appear on Q1 rows,
                    # so remember the latest non-empty values.

                    if ml_data.get("Topic"):
                        ml_current_topic = ml_data["Topic"]

                    if ml_data.get("Sub_topic"):
                        ml_current_sub_topic = ml_data["Sub_topic"]

                    if ml_data.get("Scenario"):
                        ml_current_scenario = ml_data["Scenario"]

                    if ml_data.get("Scenario_explanation"):
                        ml_current_scenario_explanation = (
                            ml_data["Scenario_explanation"]
                        )

                    if ml_data.get("Memory_shortcut"):
                        ml_current_memory_shortcut = (
                            ml_data["Memory_shortcut"]
                        )

                    if ml_data.get("Applies_when"):
                        ml_current_applies_when = (
                            ml_data["Applies_when"]
                        )

                    if ml_data.get("Not_applies_when"):
                        ml_current_not_applies_when = (
                            ml_data["Not_applies_when"]
                        )

                    # Store Malayalam data by Practice_question_id
                    malayalam_lookup[str(practice_id).strip()] = {

                        "topic_ml": ml_current_topic,
                        "sub_topic_ml": ml_current_sub_topic,

                        "scenario_ml": ml_current_scenario,

                        "scenario_explanation_ml":
                            ml_current_scenario_explanation,

                        "memory_shortcut_ml":
                            ml_current_memory_shortcut,

                        "applies_when_ml":
                            ml_current_applies_when,

                        "not_applies_when_ml":
                            ml_current_not_applies_when,

                        "question_ml":
                            ml_data.get("Questions"),

                        "option_a_ml":
                            ml_data.get("Option_A"),

                        "option_b_ml":
                            ml_data.get("Option_B"),

                        "option_c_ml":
                            ml_data.get("Option_C"),

                        "correct_answer_ml":
                            ml_data.get("Correct_answer"),

                        "explanation_ml":
                            ml_data.get("Explanation"),

                        "wrong_answer_tip_ml":
                            ml_data.get("Wrong_Answer_Tip")
                    }

                print(
                    "Malayalam rows loaded:",
                    len(malayalam_lookup)
                )

            except Exception as e:

                print(
                    "Failed to read Malayalam workbook:",
                    e
                )

                flash(
                    "Could not read the Malayalam Excel file.",
                    "danger"
                )

                return redirect(request.url)
        
        
        # Read column headers (Excel Row 1)
        headers = next(
            sheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )
        
        
        # ==========================================
        # VALIDATE REQUIRED COLUMNS
        # ==========================================

        required_columns = {
            "Topic_id",
            "Topic",
            "Sub_topic",
            "Scenario",
            "Concept_image_ref",
            "Scenario_explanation",
            "Memory_shortcut",
            "Applies_when",
            "Not_applies_when",
            "Practice_question_id",
            "Questions",
            "Option_A",
            "Option_B",
            "Option_C",
            "Correct_answer_letter",
            "Correct_answer",
            "Explanation",
            "Wrong_Answer_Tip",
            "Question_image_ref"
        }

        actual_headers = {
            header
            for header in headers
            if header is not None
        }

        missing_columns = (
            required_columns - actual_headers
        )

        if missing_columns:

            flash(
                "Invalid English dataset. Missing columns: "
                + ", ".join(sorted(missing_columns)),
                "danger"
            )

            return redirect(request.url)
        
        
        
        # ==========================================
        # CHECK CURRENT PROJECT DATA
        # ==========================================

        existing_count = DatasetEntry.query.filter_by(
            project_id=project.id
        ).count()


        # A "new" import is only valid when empty
        if (
            existing_action == "new"
            and existing_count > 0
        ):

            flash(
                "This project already contains data. "
                "Please choose Merge or Replace.",
                "warning"
            )

            return redirect(request.url)


        # Merge requires existing data
        if (
            existing_action == "merge"
            and existing_count == 0
        ):

            existing_action = "new"


        # Replace is allowed, but deletion happens
        # only AFTER the uploaded file was validated.
# Replace is allowed, but deletion happens
# only AFTER the uploaded file was validated.
        if existing_action == "replace":

            # Get IDs of existing dataset entries
            existing_entry_ids = [
                entry.id
                for entry in DatasetEntry.query.filter_by(
                    project_id=project.id
                ).all()
            ]

            # Delete corrections connected to those entries FIRST
            if existing_entry_ids:

                FieldCorrection.query.filter(
                    FieldCorrection.entry_id.in_(
                        existing_entry_ids
                    )
                ).delete(
                    synchronize_session=False
                )

            # Now delete the dataset entries
            DatasetEntry.query.filter_by(
                project_id=project.id
            ).delete(
                synchronize_session=False
            )

            # IMPORTANT:
            # Still do NOT commit here.
            # Your commit at the end of the import handles everything.

            print(
                "Existing entries and corrections marked for replacement."
            )

        # ==========================================
        # PRELOAD EXISTING ENTRIES
        # Avoid one database query per Excel row
        # ==========================================

        existing_entries = {}

        if existing_action == "merge":

            project_entries = DatasetEntry.query.filter_by(
                project_id=project.id
            ).all()

            existing_entries = {
                str(item.practice_question_id).strip(): item
                for item in project_entries
                if item.practice_question_id
            }

            print(
                "Existing entries loaded for merge:",
                len(existing_entries)
            )
        
        
        
        # Remember the latest non-empty values
        current_topic_id = None
        current_topic = None
        current_sub_topic = None

        current_scenario = None
        current_concept_image_ref = None
        current_scenario_explanation = None
        current_memory_shortcut = None
        current_applies_when = None
        current_not_applies_when = None
        
        
        
        
       # Read every data row
        for excel_row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):

            row_data = {
                header: value
                for header, value in zip(headers, row)
                if header is not None
            }


            
            # Skip completely empty rows
            if row_data.get("Practice_question_id") is None:
                continue
            
            practice_id = str(
                row_data["Practice_question_id"]
            ).strip()

            ml_data = malayalam_lookup.get(
                practice_id,
                {}
            )
                        
            
            # Remember latest values if they are not empty

            if row_data["Topic_id"]:
                current_topic_id = row_data["Topic_id"]

            if row_data["Topic"]:
                current_topic = row_data["Topic"]

            if row_data["Sub_topic"]:
                current_sub_topic = row_data["Sub_topic"]

            if row_data["Scenario"]:
                current_scenario = row_data["Scenario"]

            if row_data["Concept_image_ref"]:
                current_concept_image_ref = row_data["Concept_image_ref"]

            if row_data["Scenario_explanation"]:
                current_scenario_explanation = row_data["Scenario_explanation"]

            if row_data["Memory_shortcut"]:
                current_memory_shortcut = row_data["Memory_shortcut"]

            if row_data["Applies_when"]:
                current_applies_when = row_data["Applies_when"]

            if row_data["Not_applies_when"]:
                current_not_applies_when = row_data["Not_applies_when"]
                
                
        
            entry = DatasetEntry(
                project_id=project.id,
                row_number=excel_row_number,

                topic_id=current_topic_id,
                practice_question_id=row_data["Practice_question_id"],

                topic_en=current_topic,
                topic_ml=ml_data.get("topic_ml"),

                sub_topic_en=current_sub_topic,
                sub_topic_ml=ml_data.get("sub_topic_ml"),


                scenario_en=current_scenario,
                scenario_explanation_en=current_scenario_explanation,
                memory_shortcut_en=current_memory_shortcut,
                applies_when_en=current_applies_when,
                not_applies_when_en=current_not_applies_when,
                
                scenario_ml=ml_data.get("scenario_ml"),

                scenario_explanation_ml=ml_data.get(
                    "scenario_explanation_ml"
                ),

                memory_shortcut_ml=ml_data.get(
                    "memory_shortcut_ml"
                ),

                applies_when_ml=ml_data.get(
                    "applies_when_ml"
                ),

                not_applies_when_ml=ml_data.get(
                    "not_applies_when_ml"
                ),
                

                question_en=row_data["Questions"],
                question_ml=ml_data.get("question_ml"),
                option_a_en=row_data["Option_A"],
                option_b_en=row_data["Option_B"],
                option_c_en=row_data["Option_C"],
                option_a_ml=ml_data.get("option_a_ml"),
                option_b_ml=ml_data.get("option_b_ml"),
                option_c_ml=ml_data.get("option_c_ml"),

                correct_answer_letter=row_data["Correct_answer_letter"],
                correct_answer_en=row_data["Correct_answer"],
                correct_answer_ml=ml_data.get(
                    "correct_answer_ml"
                ),

                explanation_en=row_data["Explanation"],
                wrong_answer_tip_en=row_data["Wrong_Answer_Tip"],
                explanation_ml=ml_data.get(
                    "explanation_ml"
                ),

                wrong_answer_tip_ml=ml_data.get(
                    "wrong_answer_tip_ml"
                ),

                question_image_ref=row_data["Question_image_ref"],
                english_completed=False,
                malayalam_completed=False,
                attempted=False,
                has_correction=False
            )
            
            # ======================================
            # SAVE / MERGE ENTRY
            # ======================================

            if existing_action == "merge":
                existing_entry = existing_entries.get(practice_id)
            else:
                existing_entry = None


            if existing_entry:
                
                    # Track whether this import added any new information
                data_added = False

        # ======================================
        # SAFE MERGE
        # Preserve existing DMS data.
        # Only fill fields that are currently empty.
        # ======================================
        


                # ======================================
                # SAFE MERGE - ENGLISH
                # Track if ANY new data was added
                # ======================================

                if not existing_entry.topic_en and current_topic:
                    existing_entry.topic_en = current_topic
                    data_added = True

                if not existing_entry.sub_topic_en and current_sub_topic:
                    existing_entry.sub_topic_en = current_sub_topic
                    data_added = True

                if not existing_entry.scenario_en and current_scenario:
                    existing_entry.scenario_en = current_scenario
                    data_added = True

                if (
                    not existing_entry.scenario_explanation_en
                    and current_scenario_explanation
                ):
                    existing_entry.scenario_explanation_en = (
                        current_scenario_explanation
                    )
                    data_added = True

                if (
                    not existing_entry.memory_shortcut_en
                    and current_memory_shortcut
                ):
                    existing_entry.memory_shortcut_en = (
                        current_memory_shortcut
                    )
                    data_added = True

                if (
                    not existing_entry.applies_when_en
                    and current_applies_when
                ):
                    existing_entry.applies_when_en = current_applies_when
                    data_added = True

                if (
                    not existing_entry.not_applies_when_en
                    and current_not_applies_when
                ):
                    existing_entry.not_applies_when_en = (
                        current_not_applies_when
                    )
                    data_added = True

                if (
                    not existing_entry.question_en
                    and row_data.get("Questions")
                ):
                    existing_entry.question_en = row_data["Questions"]
                    data_added = True

                if (
                    not existing_entry.option_a_en
                    and row_data.get("Option_A")
                ):
                    existing_entry.option_a_en = row_data["Option_A"]
                    data_added = True

                if (
                    not existing_entry.option_b_en
                    and row_data.get("Option_B")
                ):
                    existing_entry.option_b_en = row_data["Option_B"]
                    data_added = True

                if (
                    not existing_entry.option_c_en
                    and row_data.get("Option_C")
                ):
                    existing_entry.option_c_en = row_data["Option_C"]
                    data_added = True

                if (
                    not existing_entry.correct_answer_letter
                    and row_data.get("Correct_answer_letter")
                ):
                    existing_entry.correct_answer_letter = (
                        row_data["Correct_answer_letter"]
                    )
                    data_added = True

                if (
                    not existing_entry.correct_answer_en
                    and row_data.get("Correct_answer")
                ):
                    existing_entry.correct_answer_en = (
                        row_data["Correct_answer"]
                    )
                    data_added = True

                if (
                    not existing_entry.explanation_en
                    and row_data.get("Explanation")
                ):
                    existing_entry.explanation_en = (
                        row_data["Explanation"]
                    )
                    data_added = True

                if (
                    not existing_entry.wrong_answer_tip_en
                    and row_data.get("Wrong_Answer_Tip")
                ):
                    existing_entry.wrong_answer_tip_en = (
                        row_data["Wrong_Answer_Tip"]
                    )
                    data_added = True

                    
                    
                # ======================================
                # SAFE MERGE - MALAYALAM
                # Only fill missing Malayalam values
                # ======================================

                if ml_data:


                    if not existing_entry.topic_ml and ml_data.get("topic_ml"):
                        existing_entry.topic_ml = ml_data["topic_ml"]
                        data_added = True

                    if not existing_entry.sub_topic_ml and ml_data.get("sub_topic_ml"):
                        existing_entry.sub_topic_ml = ml_data["sub_topic_ml"]
                        data_added = True

                    if not existing_entry.scenario_ml and ml_data.get("scenario_ml"):
                        existing_entry.scenario_ml = ml_data["scenario_ml"]
                        data_added = True
                        

                    if (
                        not existing_entry.scenario_explanation_ml
                        and ml_data.get("scenario_explanation_ml")
                    ):
                        existing_entry.scenario_explanation_ml = (
                            ml_data["scenario_explanation_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.memory_shortcut_ml
                        and ml_data.get("memory_shortcut_ml")
                    ):
                        existing_entry.memory_shortcut_ml = (
                            ml_data["memory_shortcut_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.applies_when_ml
                        and ml_data.get("applies_when_ml")
                    ):
                        existing_entry.applies_when_ml = (
                            ml_data["applies_when_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.not_applies_when_ml
                        and ml_data.get("not_applies_when_ml")
                    ):
                        existing_entry.not_applies_when_ml = (
                            ml_data["not_applies_when_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.question_ml
                        and ml_data.get("question_ml")
                    ):
                        existing_entry.question_ml = (
                            ml_data["question_ml"]
                        )
                        data_added = True
                        

                    if (
                        not existing_entry.option_a_ml
                        and ml_data.get("option_a_ml")
                    ):
                        existing_entry.option_a_ml = (
                            ml_data["option_a_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.option_b_ml
                        and ml_data.get("option_b_ml")
                    ):
                        existing_entry.option_b_ml = (
                            ml_data["option_b_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.option_c_ml
                        and ml_data.get("option_c_ml")
                    ):
                        existing_entry.option_c_ml = (
                            ml_data["option_c_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.correct_answer_ml
                        and ml_data.get("correct_answer_ml")
                    ):
                        existing_entry.correct_answer_ml = (
                            ml_data["correct_answer_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.explanation_ml
                        and ml_data.get("explanation_ml")
                    ):
                        existing_entry.explanation_ml = (
                            ml_data["explanation_ml"]
                        )
                        data_added = True

                    if (
                        not existing_entry.wrong_answer_tip_ml
                        and ml_data.get("wrong_answer_tip_ml")
                    ):
                        existing_entry.wrong_answer_tip_ml = (
                            ml_data["wrong_answer_tip_ml"]
                        )
                        data_added = True
                        
                        
                    # Mark existing entry as NEW if merge filled any missing field
                if data_added:
                    existing_entry.is_new = True
                    




            else:

                # Only rows newly introduced through MERGE
                # should receive the NEW badge.
                if existing_action == "merge":
                    entry.is_new = True
                else:
                    entry.is_new = False

                db.session.add(entry)


        print("All rows imported successfully!")

 
        

        total_rows = sheet.max_row
        total_columns = sheet.max_column

        print("=" * 50)
        print("Workbook Loaded Successfully")
        print("Sheet Name:", sheet.title)
        print("Rows:", total_rows)
        print("Columns:", total_columns)
        print("=" * 50)

        project.template_excel_path = english_path
        project.status = "Dataset Imported"

        try:
            db.session.commit()
            print("\nDATABASE COMMIT SUCCESSFUL")


        except Exception as e:
            db.session.rollback()

            print("IMPORT FAILED:", e)

            flash(
                "Import failed. No database changes were saved.",
                "danger"
            )

            return redirect(request.url)

        flash(
            "Dataset uploaded successfully!",
            "success"
        )

        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project.id
            )
        )

    existing_count = DatasetEntry.query.filter_by(
        project_id=project.id
    ).count()

    return render_template(
        "import_dataset.html",
        project=project,
        existing_count=existing_count
    )
    
    
    
@project.route("/project/<int:project_id>/export/<language>")
@login_required

def export_dataset(project_id, language):

    project = Project.query.get_or_404(project_id)

    # ---------------------------------
    # Export mode
    # ---------------------------------

    export_mode = request.args.get("mode", "final")

    query = DatasetEntry.query.filter_by(
        project_id=project.id
    )

    if export_mode == "final":

        # Final dataset = Approved entries only
        query = query.filter(
            DatasetEntry.status == DatasetEntry.STATUS_APPROVED
        )

    elif export_mode == "custom":

        selected_statuses = request.args.getlist("status")

        if not selected_statuses:
            flash("Please select at least one status to export.", "warning")

            return redirect(
                url_for(
                    "project.project_dashboard",
                    project_id=project.id
                )
            )

        query = query.filter(
            DatasetEntry.status.in_(selected_statuses)
        )

    else:

        flash("Invalid export mode.", "danger")

        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project.id
            )
        )

    entries = query.order_by(
        DatasetEntry.row_number
    ).all()

    if not entries:

        if export_mode == "final":
            message = "No approved entries available for final export."
        else:
            message = "No entries found for the selected statuses."

        flash(message, "warning")

        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project.id
            )
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dataset"

    headers = [

    "Topic_id",

    "Topic",

    "Sub_topic",

    "Scenario",

    "Concept_image_ref",

    "Scenario_explanation",

    "Memory_shortcut",

    "Applies_when",

    "Not_applies_when",

    "Practice_question_id",

    "Questions",

    "Option_A",

    "Option_B",

    "Option_C",

    "Correct_answer_letter",

    "Correct_answer",

    "Explanation",

    "Wrong_Answer_Tip",

    "Question_image_ref"

]
    
    if export_mode == "custom":
        headers.append("Status")

    sheet.append(headers)
    
    
    if language == "english":

        topic = "topic_en"
        sub_topic = "sub_topic_en"

        scenario = "scenario_en"
        scenario_explanation = "scenario_explanation_en"

        memory_shortcut = "memory_shortcut_en"
        applies_when = "applies_when_en"
        not_applies_when = "not_applies_when_en"

        question = "question_en"

        option_a = "option_a_en"
        option_b = "option_b_en"
        option_c = "option_c_en"

        correct_answer = "correct_answer_en"

        explanation = "explanation_en"

        wrong_answer_tip = "wrong_answer_tip_en"

    elif language == "malayalam":

        topic = "topic_ml"
        sub_topic = "sub_topic_ml"

        scenario = "scenario_ml"
        scenario_explanation = "scenario_explanation_ml"

        memory_shortcut = "memory_shortcut_ml"
        applies_when = "applies_when_ml"
        not_applies_when = "not_applies_when_ml"

        question = "question_ml"

        option_a = "option_a_ml"
        option_b = "option_b_ml"
        option_c = "option_c_ml"

        correct_answer = "correct_answer_ml"

        explanation = "explanation_ml"

        wrong_answer_tip = "wrong_answer_tip_ml"

    else:

        flash("Invalid export type", "danger")

        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project.id
            )
        )
    
    
    previous_topic = None


    for entry in entries:

        if entry.topic_id != previous_topic:

            topic_id = entry.topic_id

            topic_value = getattr(entry, topic)
            sub_topic_value = getattr(entry, sub_topic)

            scenario_value = getattr(entry, scenario)

            concept_image_ref = entry.topic_id.lower()

            scenario_explanation_value = getattr(
                entry,
                scenario_explanation
            )

            memory_shortcut_value = getattr(
                entry,
                memory_shortcut
            )

            applies_when_value = getattr(
                entry,
                applies_when
            )

            not_applies_when_value = getattr(
                entry,
                not_applies_when
            )

            previous_topic = entry.topic_id

        else:

            topic_id = ""
            topic_value = ""
            sub_topic_value = ""
            scenario_value = ""
            concept_image_ref = ""
            scenario_explanation_value = ""
            memory_shortcut_value = ""
            applies_when_value = ""
            not_applies_when_value = ""

        # Every entry gets exported
        row = [

            topic_id,
            topic_value,
            sub_topic_value,
            scenario_value,
            concept_image_ref,
            scenario_explanation_value,
            memory_shortcut_value,
            applies_when_value,
            not_applies_when_value,

            entry.practice_question_id,

            getattr(entry, question),

            getattr(entry, option_a),
            getattr(entry, option_b),
            getattr(entry, option_c),

            entry.correct_answer_letter,

            getattr(entry, correct_answer),

            getattr(entry, explanation),

            getattr(entry, wrong_answer_tip),

            entry.question_image_ref

        ]

        if export_mode == "custom":
            row.append(entry.status)

        sheet.append(row)


    from io import BytesIO

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            f"{project.project_name}_{language}_"
            f"{'final' if export_mode == 'final' else 'custom'}.xlsx"
        ),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        
    
    
@project.route("/project/<int:project_id>/corrections")
def corrections(project_id):

    project = Project.query.get_or_404(project_id)

    entries = DatasetEntry.query.filter(

        DatasetEntry.project_id == project.id,

        DatasetEntry.correction_note.isnot(None),

        DatasetEntry.correction_note != ""

    ).all()

    return render_template(
        "corrections.html",
        project=project,
        entries=entries
    )
    
    
    
@project.route("/project/<int:project_id>/continue")
def continue_editing(project_id):

    project = Project.query.get_or_404(project_id)

    entry = DatasetEntry.query.filter_by(
        project_id=project.id,
        attempted=False
    ).order_by(DatasetEntry.row_number).first()

    if not entry:

        flash("🎉 All questions have been reviewed!", "success")

        return redirect(
            url_for(
                "project.view_dataset",
                project_id=project.id
            )
        )

    return redirect(
        url_for(
            "project.edit_question",
            entry_id=entry.id
        )
    )
    
    
    
@project.route("/project/<int:project_id>/new-entry", methods=["GET", "POST"])
def new_entry(project_id):

    project = Project.query.get_or_404(project_id)

    if request.method == "POST":

        entry = DatasetEntry()

        entry.project_id = project.id
        
        selected_topic = request.form.get("topic_en")

        if selected_topic == "__NEW__":
            entry.topic_en = request.form.get("new_topic_name")
        else:
            entry.topic_en = selected_topic
        entry.topic_ml = request.form.get("topic_ml")

        selected_subtopic = request.form.get("sub_topic_en")

        if selected_subtopic == "__NEW__":
            entry.sub_topic_en = request.form.get("new_sub_topic")
        else:
            entry.sub_topic_en = selected_subtopic
        entry.sub_topic_ml = request.form.get("sub_topic_ml")

        # -------------------------------
        # Scenario
        # -------------------------------
        entry.scenario_en = request.form.get("scenario_en")
        entry.scenario_ml = request.form.get("scenario_ml")

        entry.scenario_explanation_en = request.form.get("scenario_explanation_en")
        entry.scenario_explanation_ml = request.form.get("scenario_explanation_ml")

        entry.memory_shortcut_en = request.form.get("memory_shortcut_en")
        entry.memory_shortcut_ml = request.form.get("memory_shortcut_ml")

        entry.applies_when_en = request.form.get("applies_when_en")
        entry.applies_when_ml = request.form.get("applies_when_ml")

        entry.not_applies_when_en = request.form.get("not_applies_when_en")
        entry.not_applies_when_ml = request.form.get("not_applies_when_ml")

        # -------------------------------
        # Question
        # -------------------------------
        entry.question_en = request.form.get("question_en")
        entry.question_ml = request.form.get("question_ml")

        entry.option_a_en = request.form.get("option_a_en")
        entry.option_a_ml = request.form.get("option_a_ml")

        entry.option_b_en = request.form.get("option_b_en")
        entry.option_b_ml = request.form.get("option_b_ml")

        entry.option_c_en = request.form.get("option_c_en")
        entry.option_c_ml = request.form.get("option_c_ml")

        entry.correct_answer_letter = request.form.get("correct_answer_letter")

        entry.correct_answer_en = request.form.get("correct_answer_en")
        entry.correct_answer_ml = request.form.get("correct_answer_ml")

        entry.explanation_en = request.form.get("explanation_en")
        entry.explanation_ml = request.form.get("explanation_ml")

        entry.wrong_answer_tip_en = request.form.get("wrong_answer_tip_en")
        entry.wrong_answer_tip_ml = request.form.get("wrong_answer_tip_ml")

    # -------------------------------
    # Other Fields
    # -------------------------------
        entry.question_image_ref = request.form.get("question_image_ref")

        entry.correction_note = request.form.get("correction_note")

        entry.english_completed = "english_completed" in request.form
        entry.malayalam_completed = "malayalam_completed" in request.form

        entry.attempted = True
        
        
        
        last_entry = DatasetEntry.query.filter_by(project_id=project.id) \
            .order_by(DatasetEntry.row_number.desc()) \
            .first()

        if last_entry:
            entry.row_number = last_entry.row_number + 1
        else:
            entry.row_number = 1
            
            
        selected_topic = entry.topic_en

        topics = DatasetEntry.query.filter_by(
            project_id=project.id,
            topic_en=selected_topic
        ).all()

        if topics:

            prefix = topics[0].topic_id.split("_")[0]

            max_number = max(
                int(t.topic_id.split("_")[1])
                for t in topics
            )

            entry.topic_id = f"{prefix}_{max_number + 1}"

        else:
            topic_code = request.form.get("topic_code", "").upper()
            entry.topic_id = f"{topic_code}_1"

        entry.practice_question_id = f"{entry.topic_id}_Q1"

        db.session.add(entry)

        try:
            db.session.commit()

        except Exception as e:

            db.session.rollback()

            print(
                f"NEW ENTRY SAVE FAILED | "
                f"Project ID: {project.id} | "
                f"Error: {repr(e)}"
            )

            flash(
                "The new entry could not be saved due to a temporary "
                "server or database issue. Please try again.",
                "danger"
            )

            return redirect(
                url_for(
                    "project.new_entry",
                    project_id=project.id
                )
            )

        flash("New dataset entry created successfully!", "success")

        return redirect(
            url_for(
                "project.edit_question",
                entry_id=entry.id,
                saved=1,
                clear_new_draft=project.id
            )
        )
        
    topics = (
        db.session.query(DatasetEntry.topic_en)
        .distinct()
        .order_by(DatasetEntry.topic_en)
        .all()
    )

    topics = [t[0] for t in topics if t[0]]
    
    
    topic_translations = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.topic_ml
        )
        .distinct()
        .all()
    )

    subtopic_translations = (
        db.session.query(
            DatasetEntry.sub_topic_en,
            DatasetEntry.sub_topic_ml
        )
        .distinct()
        .all()
    )
    

    subtopics = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .distinct()
        .order_by(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .all()
    )
    
    subtopic_details = {}

    entries = DatasetEntry.query.filter_by(
        project_id=project.id
    ).all()

    for e in entries:

        key = f"{e.topic_en}|{e.sub_topic_en}"

        if key not in subtopic_details:

            subtopic_details[key] = {

                "scenario_en": e.scenario_en,
                "scenario_ml": e.scenario_ml,

                "scenario_explanation_en": e.scenario_explanation_en,
                "scenario_explanation_ml": e.scenario_explanation_ml,

                "memory_shortcut_en": e.memory_shortcut_en,
                "memory_shortcut_ml": e.memory_shortcut_ml,

                "applies_when_en": e.applies_when_en,
                "applies_when_ml": e.applies_when_ml,

                "not_applies_when_en": e.not_applies_when_en,
                "not_applies_when_ml": e.not_applies_when_ml
            }
            
            
    
    return render_template(
        "edit_question.html",
        project=project,
        entry=None,
        previous_entry=None,
        next_entry=None,
        topics=topics,
        subtopics=subtopics,
        topic_translations=topic_translations,
        subtopic_translations=subtopic_translations,
        correction_map={},
        correction_count=0,
        role=session.get("role"),
        subtopic_details=subtopic_details,
        shared_fields_locked=False,
        shared_entry=None,
    )
    
    
    
    
@project.route("/project/<int:project_id>/review-queue/<status>")
@login_required
def review_queue(project_id, status):

    project = Project.query.get_or_404(project_id)

    status_map = {
        "draft": DatasetEntry.STATUS_DRAFT,
        "submitted": DatasetEntry.STATUS_SUBMITTED,
        "corrections": DatasetEntry.STATUS_CORRECTIONS,
        "approved": DatasetEntry.STATUS_APPROVED
    }

    if status not in status_map:
        flash("Invalid status.", "danger")
        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project.id
            )
        )

    entries = DatasetEntry.query.filter_by(
        project_id=project.id,
        status=status_map[status]
    ).order_by(
        DatasetEntry.row_number
    ).all()

    return render_template(
        "review_queue.html",
        project=project,
        entries=entries,
        current_status=status
    )
    
    
    
    
@project.route("/correction/save", methods=["POST"])
@login_required
def save_field_correction():

    data = request.get_json()

    entry_id = data.get("entry_id")
    field_name = data.get("field_name")
    comment = data.get("comment")

    correction = FieldCorrection.query.filter_by(
        entry_id=entry_id,
        field_name=field_name,
        is_active=True
    ).first()

    if correction:

        correction.comment = comment

    else:

        correction = FieldCorrection(
            entry_id=entry_id,
            field_name=field_name,
            comment=comment,
            created_by=session["user_id"]
        )

        db.session.add(correction)

        try:
            db.session.commit()

        except Exception as e:
            db.session.rollback()

            print("ERROR saving field correction:", e)

            return jsonify({
                "success": False,
                "message": "Could not save correction. Please try again."
            }), 500


        return jsonify({
            "success": True
        })
    
    
@project.route("/correction/<int:entry_id>")
@login_required
def get_field_corrections(entry_id):

    corrections = FieldCorrection.query.filter_by(
        entry_id=entry_id,
        is_active=True
    ).all()

    result = {}

    for correction in corrections:
        result[correction.field_name] = correction.comment

    return jsonify(result)
    
@project.route(
    "/project/<int:project_id>/review/<int:entry_id>",
    methods=["GET", "POST"]
)
@login_required
def review_question(project_id, entry_id):

    project = Project.query.get_or_404(project_id)

    entry = DatasetEntry.query.get_or_404(entry_id)
    
    shared_fields_locked = (
        entry.practice_question_id
        and not entry.practice_question_id.endswith("_Q1")
    )
    
    if request.method == "POST":

        action = request.form.get("action")
        
        if entry.status != DatasetEntry.STATUS_SUBMITTED:

            flash(
                "This question is not currently available for review.",
                "warning"
            )

            return redirect(
                url_for(
                    "project.review_question",
                    project_id=project.id,
                    entry_id=entry.id
                )
            )
                
        
        
        # Shared Fields
        # Only Q1 can update these
        # ---------------------------------

        if not shared_fields_locked:

            entry.topic_en = request.form.get("topic_en")
            entry.topic_ml = request.form.get("topic_ml")

            entry.sub_topic_en = request.form.get("sub_topic_en")
            entry.sub_topic_ml = request.form.get("sub_topic_ml")

            entry.scenario_en = request.form.get("scenario_en")
            entry.scenario_ml = request.form.get("scenario_ml")

            entry.scenario_explanation_en = request.form.get(
                "scenario_explanation_en"
            )
            entry.scenario_explanation_ml = request.form.get(
                "scenario_explanation_ml"
            )

            entry.memory_shortcut_en = request.form.get(
                "memory_shortcut_en"
            )
            entry.memory_shortcut_ml = request.form.get(
                "memory_shortcut_ml"
            )

            entry.applies_when_en = request.form.get(
                "applies_when_en"
            )
            entry.applies_when_ml = request.form.get(
                "applies_when_ml"
            )

            entry.not_applies_when_en = request.form.get(
                "not_applies_when_en"
            )
            entry.not_applies_when_ml = request.form.get(
                "not_applies_when_ml"
            )

        entry.question_en = request.form.get("question_en")
        entry.question_ml = request.form.get("question_ml")

        entry.option_a_en = request.form.get("option_a_en")
        entry.option_a_ml = request.form.get("option_a_ml")

        entry.option_b_en = request.form.get("option_b_en")
        entry.option_b_ml = request.form.get("option_b_ml")

        entry.option_c_en = request.form.get("option_c_en")
        entry.option_c_ml = request.form.get("option_c_ml")

        entry.correct_answer_en = request.form.get("correct_answer_en")
        entry.correct_answer_ml = request.form.get("correct_answer_ml")

        entry.correct_answer_letter = request.form.get("correct_answer_letter")

        entry.explanation_en = request.form.get("explanation_en")
        entry.explanation_ml = request.form.get("explanation_ml")

        entry.wrong_answer_tip_en = request.form.get("wrong_answer_tip_en")
        entry.wrong_answer_tip_ml = request.form.get("wrong_answer_tip_ml")

        entry.question_image_ref = request.form.get("question_image_ref")
        
        # Save DMS-only rich formatting
        # ---------------------------------

        rich_content = dict(entry.rich_content or {})

        shared_field_names = {
            "topic_en",
            "topic_ml",
            "sub_topic_en",
            "sub_topic_ml",
            "scenario_en",
            "scenario_ml",
            "scenario_explanation_en",
            "scenario_explanation_ml",
            "memory_shortcut_en",
            "memory_shortcut_ml",
            "applies_when_en",
            "applies_when_ml",
            "not_applies_when_en",
            "not_applies_when_ml",
        }

        for key, value in request.form.items():

            if not key.startswith("rich_"):
                continue

            field_name = key[5:]

            # Q2+ cannot change shared formatting
            if shared_fields_locked and field_name in shared_field_names:
                continue

            if value:
                rich_content[field_name] = value
            else:
                rich_content.pop(field_name, None)

        entry.rich_content = rich_content

        # ---------------------------------
        # APPROVE / REQUEST CORRECTIONS
        # Then move to next Pending Review
        # ---------------------------------

        # Find the next submitted entry BEFORE changing current entry's status
        next_pending_entry = (
            DatasetEntry.query
            .filter(
                DatasetEntry.project_id == project.id,
                DatasetEntry.status == DatasetEntry.STATUS_SUBMITTED,
                DatasetEntry.id > entry.id
            )
            .order_by(DatasetEntry.id.asc())
            .first()
        )

        # If there is nothing after this entry, look from the beginning.
        # This prevents getting sent back to the queue while other pending
        # entries with smaller IDs still exist.
        if not next_pending_entry:
            next_pending_entry = (
                DatasetEntry.query
                .filter(
                    DatasetEntry.project_id == project.id,
                    DatasetEntry.status == DatasetEntry.STATUS_SUBMITTED,
                    DatasetEntry.id != entry.id
                )
                .order_by(DatasetEntry.id.asc())
                .first()
            )


        # ---------------------------------
        # Find the next Pending Review
        # ---------------------------------

        next_pending_entry = (
            DatasetEntry.query
            .filter(
                DatasetEntry.project_id == project.id,
                DatasetEntry.status == DatasetEntry.STATUS_SUBMITTED,
                DatasetEntry.row_number > entry.row_number
            )
            .order_by(DatasetEntry.row_number.asc())
            .first()
        )

        # If there is no pending entry after this one,
        # check whether another pending entry exists earlier
        if not next_pending_entry:
            next_pending_entry = (
                DatasetEntry.query
                .filter(
                    DatasetEntry.project_id == project.id,
                    DatasetEntry.status == DatasetEntry.STATUS_SUBMITTED,
                    DatasetEntry.id != entry.id
                )
                .order_by(DatasetEntry.row_number.asc())
                .first()
            )


        # ---------------------------------
        # APPROVE
        # ---------------------------------

        if action == "approve":

            entry.status = DatasetEntry.STATUS_APPROVED
            entry.approved_at = datetime.utcnow()

            # Deactivate reviewer comments after approval
            FieldCorrection.query.filter_by(
                entry_id=entry.id,
                is_active=True
            ).update({
                "is_active": False
            })

            success_message = "Entry approved successfully!"


        # ---------------------------------
        # REQUEST CORRECTIONS
        # ---------------------------------

        elif action == "corrections":

            entry.status = DatasetEntry.STATUS_CORRECTIONS
            entry.approved_at = None

            success_message = "Corrections requested successfully!"


        # ---------------------------------
        # INVALID ACTION
        # ---------------------------------

        else:

            flash("Invalid review action.", "danger")

            return redirect(
                url_for(
                    "project.review_question",
                    project_id=project.id,
                    entry_id=entry.id,
                    from_status="submitted"
                )
            )


        # Save changes
        db.session.commit()

        flash(success_message, "success")


        # ---------------------------------
        # Open next Pending Review
        # ---------------------------------

        if next_pending_entry:

            return redirect(
                url_for(
                    "project.review_question",
                    project_id=project.id,
                    entry_id=next_pending_entry.id,
                    from_status="submitted"
                )
            )


        # ---------------------------------
        # No Pending Reviews left
        # Return to Pending Reviews queue
        # ---------------------------------

        return redirect(
            url_for(
                "project.review_queue",
                project_id=project.id,
                status="submitted"
            )
        )

        db.session.commit()

        flash(success_message, "success")


        # ---------------------------------
        # Move reviewer to next Pending Review
        # ---------------------------------

        if next_pending_entry:

            return redirect(
                url_for(
                    "project.review_question",
                    project_id=project.id,
                    entry_id=next_pending_entry.id,
                    from_status="submitted"
                )
            )


        # No Pending Reviews remaining
        return redirect(
            url_for(
                "project.review_queue",
                project_id=project.id,
                status="submitted"
            )
        )
    
    
    role = session.get("role")

    reviewer_editing_locked = (
        entry.status != DatasetEntry.STATUS_SUBMITTED
    )

    corrections = FieldCorrection.query.filter_by(
        entry_id=entry.id,
        is_active=True
    ).all()

    correction_map = {
        c.field_name: c.comment
        for c in corrections
    }
    
    correction_count = len(correction_map)
    
    topics = (
    db.session.query(DatasetEntry.topic_en)
        .distinct()
        .order_by(DatasetEntry.topic_en)
        .all()
    )

    topics = [t[0] for t in topics if t[0]]

    subtopics = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .distinct()
        .order_by(
            DatasetEntry.topic_en,
            DatasetEntry.sub_topic_en
        )
        .all()
    )

    topic_translations = (
        db.session.query(
            DatasetEntry.topic_en,
            DatasetEntry.topic_ml
        )
        .distinct()
        .all()
    )

    subtopic_translations = (
        db.session.query(
            DatasetEntry.sub_topic_en,
            DatasetEntry.sub_topic_ml
        )
        .distinct()
        .all()
    )
    

    
    
    shared_entry = None

    if shared_fields_locked:

        q1_id = entry.practice_question_id.rsplit("_Q", 1)[0] + "_Q1"

        shared_entry = DatasetEntry.query.filter_by(
            project_id=entry.project_id,
            practice_question_id=q1_id
        ).first()
            
   # ---------------------------------
    # REVIEW NAVIGATION
    # Previous / Next within current queue
    # ---------------------------------

    from_status = request.args.get("from_status")

    # Normalize URL value to actual database status
    status_map = {
        "submitted": DatasetEntry.STATUS_SUBMITTED,
        "corrections": DatasetEntry.STATUS_CORRECTIONS,
        "approved": DatasetEntry.STATUS_APPROVED,
        "draft": DatasetEntry.STATUS_DRAFT,
    }

    navigation_status = status_map.get(
        from_status.lower() if from_status else None
    )

    # Base navigation query
    navigation_query = DatasetEntry.query.filter(
        DatasetEntry.project_id == project.id
    )

    # When opened from a reviewer queue, stay inside that queue
    if navigation_status:
        navigation_query = navigation_query.filter(
            DatasetEntry.status == navigation_status
        )

    previous_entry = (
        navigation_query
        .filter(DatasetEntry.id < entry.id)
        .order_by(DatasetEntry.id.desc())
        .first()
    )

    next_entry = (
        navigation_query
        .filter(DatasetEntry.id > entry.id)
        .order_by(DatasetEntry.id.asc())
        .first()
    )        
    

    return render_template(
        "review_question.html",
        project=project,
        entry=entry,
        role=role,

        topics=topics,
        subtopics=subtopics,
        topic_translations=topic_translations,
        subtopic_translations=subtopic_translations,

        correction_map=correction_map,
        correction_count=correction_count,
        subtopic_details={},
        shared_entry=shared_entry,
        shared_fields_locked=shared_fields_locked,
        previous_entry=previous_entry,
        next_entry=next_entry,
        reviewer_editing_locked=reviewer_editing_locked,
        from_status=from_status,

    )
    
@project.route("/project/<int:project_id>/team-queue/<status>")
@login_required
def team_queue(project_id, status):

    print("\n===== TEAM QUEUE CALLED =====")
    print("PROJECT ID:", project_id)
    print("STATUS FROM URL:", repr(status))
    print("SESSION ROLE:", repr(session.get("role")))

    project_obj = Project.query.get_or_404(project_id)

    role = session.get("role")

    if role not in ("TEAM", "TEAM_MEMBER"):
        print("ROLE FAILED - REDIRECTING TO DASHBOARD")

        return redirect(
            url_for(
                "project.project_dashboard",
                project_id=project_id
            )
        )

    print("ROLE PASSED")

    allowed_statuses = {
        DatasetEntry.STATUS_DRAFT,
        DatasetEntry.STATUS_SUBMITTED,
        DatasetEntry.STATUS_CORRECTIONS,
        DatasetEntry.STATUS_APPROVED,
    }

    print("ALLOWED STATUSES:", allowed_statuses)

    if status not in allowed_statuses:
        print("INVALID STATUS")
        abort(404)

    entries = (
        DatasetEntry.query
        .filter_by(
            project_id=project_id,
            status=status
        )
        .order_by(
            DatasetEntry.topic_id.asc(),
            DatasetEntry.practice_question_id.asc()
        )
        .all()
    )

    print("ENTRIES FOUND:", len(entries))

    status_labels = {
        DatasetEntry.STATUS_DRAFT: "Draft",
        DatasetEntry.STATUS_SUBMITTED: "Pending Review",
        DatasetEntry.STATUS_CORRECTIONS: "Corrections Requested",
        DatasetEntry.STATUS_APPROVED: "Approved",
    }

    current_status_label = status_labels.get(
        status,
        status.replace("_", " ").title()
    )

    print("RENDERING TEAM_QUEUE.HTML")
    print("=============================\n")

    return render_template(
        "team_queue.html",
        project=project_obj,
        entries=entries,
        current_status=status,
        current_status_label=current_status_label
    )